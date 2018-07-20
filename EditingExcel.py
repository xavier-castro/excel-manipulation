#! python3

import openpyxl

wb = openpyxl.Workbook() # creating a new workbook object
print(wb)

print(wb.sheetnames) # see what sheet names exist

sheet = wb['Sheet']
print(sheet)

sheet['A1'] = 42        # change or add data to cell
sheet['A2'] = 'Hello'
wb.save('EditingExcelSample.xlsx')  # save what you changed/added to the spreedsheat. Make sure to save to different sheet so you don't overwrite

sheet2 = wb.create_sheet() # create a new sheet
print(wb.sheetnames)
print(sheet2.title)
sheet2.title = 'My New Sheet Name'
print(wb.sheetnames)
wb.save('EditingExcelSample2.xlsx')

wb.create_sheet(index = 0, title = 'My Other Sheet')   # Creating a new sheet and saving it as the first sheet
wb.save('EditingExcelSample3.xlsx')
print(wb.sheetnames) # Confirming that it is first
